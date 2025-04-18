from flask import current_app
from flask import Blueprint
from flask import render_template, redirect, url_for, request, flash, send_file, session
from flask_login import login_user, logout_user, login_required, current_user
# from sqlalchemy.sql.functions import count
from werkzeug.security import check_password_hash
from models import User, Part
from extensions import db, login_manager
from utils import allowed_file, handle_file_upload
from sqlalchemy import or_
from openpyxl import Workbook
import io

main = Blueprint('main', __name__)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@main.route('/')
@login_required
def index():
    parts = Part.query.all()
    count = Part.query.count()
    return render_template('index.html', parts=parts, user=current_user, count=count)

@main.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('main.index'))
        flash('Invalid username or password')
    return render_template('login.html')

@main.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('main.login'))

#additional for routes

# @main.route('/add', methods=['GET', 'POST'])
# @login_required
# def add_part():
#     return render_template('add_part.html')  # пока просто заглушка

@main.route('/add', methods=['GET', 'POST'])
@login_required
def add_part():
    if request.method == 'POST':
        sap_code = request.form['sap_code']
        part_number = request.form['part_number']
        name = request.form['name']
        description = request.form['description']
        category = request.form['category']
        equipment_code = request.form['equipment_code']
        location = request.form['location']
        manufacturer = request.form['manufacturer']
        analog_group = request.form['analog_group']
        photo = request.files['photo']

        photo_path = None
        if photo and allowed_file(photo.filename):
            photo_path = handle_file_upload(photo, current_app.config['UPLOAD_FOLDER'])

        new_part = Part(
            sap_code=sap_code,
            part_number=part_number,
            name=name,
            description=description,
            category=category,
            equipment_code=equipment_code,
            location=location,
            manufacturer=manufacturer,
            analog_group=analog_group,
            photo_path=photo_path
        )

        db.session.add(new_part)
        db.session.commit()

        flash('✅ Part added successfully.')
        return redirect(url_for('main.index'))

    return render_template('add_part.html')

# @main.route('/edit/<int:part_id>', methods=['GET', 'POST'])
# @login_required
# def edit_part(part_id):
#     return render_template('edit_part.html', part_id=part_id)  # заглушка

@main.route('/edit/<int:part_id>', methods=['GET', 'POST'])
@login_required
def edit_part(part_id):
    part = Part.query.get_or_404(part_id)

    if request.method == 'POST':
        part.sap_code = request.form['sap_code']
        part.part_number = request.form['part_number']
        part.name = request.form['name']
        part.description = request.form['description']
        part.category = request.form['category']
        part.equipment_code = request.form['equipment_code']
        part.location = request.form['location']
        part.manufacturer = request.form['manufacturer']
        part.analog_group = request.form['analog_group']
        photo = request.files.get('photo')

        if photo and allowed_file(photo.filename):
            part.photo_path = handle_file_upload(photo, current_app.config['UPLOAD_FOLDER'])

        db.session.commit()
        flash('Part updated successfully.')
        return redirect(url_for('main.view_part', part_id=part.id))

    return render_template('edit_part.html', part=part, user=current_user)

# @main.route('/part/<int:part_id>')
# @login_required
# def view_part(part_id):
#     return render_template('view_part.html', part_id=part_id)  # заглушка

@main.route('/part/<int:part_id>')
@login_required
def view_part(part_id):
    part = Part.query.get_or_404(part_id)

    analogs = []
    if part.analog_group:
        analogs = Part.query.filter(
            Part.analog_group == part.analog_group,
            Part.id != part.id
        ).all()

    return render_template('view_part.html', part=part, analogs=analogs, user=current_user)

# @main.route('/delete/<int:part_id>', methods=['POST'])
# @login_required
# def delete_part(part_id):
#     flash("Delete is not implemented yet.")
#     return redirect(url_for('main.index'))  #plug

@main.route('/delete/<int:part_id>', methods=['POST'])
@login_required
def delete_part(part_id):
    if current_user.role != 'root':
        flash('❌ You do not have permission to delete parts.')
        return redirect(url_for('main.index'))

    part = Part.query.get_or_404(part_id)
    db.session.delete(part)
    db.session.commit()
    flash('✅ Part deleted successfully.')
    return redirect(url_for('main.index'))

@main.route('/export')
@login_required
def export():
    flash("Export is not implemented yet.")
    return redirect(url_for('main.index'))  #plug

# @main.route('/import')  # old version import, added not correct
# @login_required
# def import_parts():
#     flash("Import is not implemented yet.")
#     return redirect(url_for('main.index')) #plug

# @main.route('/import', methods=['GET', 'POST'])
# @login_required
# def import_parts():
#     if current_user.role not in ['admin', 'root']:
#         flash("You do not have permission to import parts.")
#         return redirect(url_for('main.index'))
#
#     if request.method == 'POST':
#         file = request.files.get('file')
#         if not file:
#             flash("No file uploaded.")
#             return redirect(url_for('main.import_parts'))
#
#         filename = file.filename.lower()
#
#         try:
#             if filename.endswith('.xlsx'):
#                 from openpyxl import load_workbook
#                 wb = load_workbook(file)
#                 ws = wb.active
#                 for row in ws.iter_rows(min_row=2, values_only=True):
#                     part = Part(
#                         sap_code=row[0],
#                         part_number=row[1],
#                         name=row[2],
#                         category=row[3],
#                         equipment_code=row[4],
#                         location=row[5],
#                         manufacturer=row[6],
#                         analog_group=row[7],
#                         description=row[8]
#                     )
#                     db.session.add(part)
#                 db.session.commit()
#
#             elif filename.endswith('.csv'):
#                 import csv, io
#                 stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
#                 reader = csv.DictReader(stream)
#                 for row in reader:
#                     part = Part(
#                         sap_code=row['sap_code'],
#                         part_number=row['part_number'],
#                         name=row['name'],
#                         category=row['category'],
#                         equipment_code=row['equipment_code'],
#                         location=row['location'],
#                         manufacturer=row['manufacturer'],
#                         analog_group=row['analog_group'],
#                         description=row['description']
#                     )
#                     db.session.add(part)
#                 db.session.commit()
#
#             else:
#                 flash("Unsupported file type. Please upload .xlsx or .csv.")
#                 return redirect(url_for('main.import_parts'))
#
#             flash("✅ Parts imported successfully!")
#             return redirect(url_for('main.index'))
#
#         except Exception as e:
#             flash(f"❌ Error during import: {e}")
#             return redirect(url_for('main.import_parts'))
#
#     return render_template('import.html')

@main.route('/import', methods=['GET', 'POST'])
@login_required
def import_parts():
    if current_user.role not in ['admin', 'root']:
        flash("You do not have permission to import parts.")
        return redirect(url_for('main.index'))

    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            flash("No file uploaded.")
            return redirect(url_for('main.import_parts'))

        filename = file.filename.lower()
        added = 0
        skipped = []

        try:
            if filename.endswith('.xlsx'):
                from openpyxl import load_workbook
                wb = load_workbook(file)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    existing = Part.query.filter_by(sap_code=row[0]).first()
                    if existing:
                        skipped.append(row[0])
                        continue
                    part = Part(
                        sap_code=row[0],
                        part_number=row[1],
                        name=row[2],
                        category=row[3],
                        equipment_code=row[4],
                        location=row[5],
                        manufacturer=row[6],
                        analog_group=row[7],
                        description=row[8]
                    )
                    db.session.add(part)
                    added += 1
                db.session.commit()

            elif filename.endswith('.csv'):
                import csv, io
                stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
                reader = csv.DictReader(stream)
                for row in reader:
                    if not row['sap_code']:
                        continue
                    existing = Part.query.filter_by(sap_code=row['sap_code']).first()
                    if existing:
                        skipped.append(row['sap_code'])
                        continue
                    part = Part(
                        sap_code=row['sap_code'],
                        part_number=row['part_number'],
                        name=row['name'],
                        category=row['category'],
                        equipment_code=row['equipment_code'],
                        location=row['location'],
                        manufacturer=row['manufacturer'],
                        analog_group=row['analog_group'],
                        description=row['description']
                    )
                    db.session.add(part)
                    added += 1
                db.session.commit()

            else:
                flash("Unsupported file type. Please upload .xlsx or .csv.")
                return redirect(url_for('main.import_parts'))

            flash(f"✅ {added} parts imported successfully.")
            if skipped:
                preview = ", ".join(skipped[:10]) + ("..." if len(skipped) > 10 else "")
                flash(f"⚠️ Skipped {len(skipped)} duplicates: {preview}")
            return redirect(url_for('main.index'))

        except Exception as e:
            flash(f"❌ Error during import: {e}")
            return redirect(url_for('main.import_parts'))

    return render_template('import.html')




@main.route('/search/results/<int:index>')
@login_required
def search_results(index):
    ids = session.get('search_results', [])
    if not ids:
        flash("No results in session.")
        return redirect(url_for('main.index'))

    if index < 0 or index >= len(ids):
        flash("Index out of range.")
        return redirect(url_for('main.search_results', index=0))

    part = Part.query.get_or_404(ids[index])
    return render_template(
        'search_results.html',
        part=part,
        index=index,
        total=len(ids),
        user=current_user
    )

from sqlalchemy import or_

@main.route('/search', methods=['GET'])
@login_required
def search():
    keyword = request.args.get('query', '').strip()

    if not keyword:
        flash("Enter a search keyword.")
        return redirect(url_for('main.index'))

    results = Part.query.filter(or_(
        Part.sap_code.ilike(f"%{keyword}%"),
        Part.part_number.ilike(f"%{keyword}%"),
        Part.name.ilike(f"%{keyword}%"),
        Part.category.ilike(f"%{keyword}%"),
        Part.equipment_code.ilike(f"%{keyword}%"),
        Part.location.ilike(f"%{keyword}%"),
        Part.manufacturer.ilike(f"%{keyword}%"),
        Part.analog_group.ilike(f"%{keyword}%"),
        Part.description.ilike(f"%{keyword}%")
    )).all()

    if not results:
        flash("No results found.")
        return redirect(url_for('main.index'))

    session['search_results'] = [p.id for p in results]
    return redirect(url_for('main.search_results', index=0))
